import qrcode
from PIL import Image, ImageDraw, ImageFont, ImageOps
from openpyxl import load_workbook

YEAR = 2022
MAIN_HEADING = 'Christmas Gift - {}'.format(YEAR)


PRIZE_LIST_FILE = '..//prize_list//PrizeList.xlsx'
PRIZE_COULUMN = 'B'
WORK_BOOK_CONTENT = load_workbook(PRIZE_LIST_FILE)
SHEET = WORK_BOOK_CONTENT.active

HEIGHT = 500
CONVERSION_FACTOR = 1.414516129032258
WIDTH = round (HEIGHT * CONVERSION_FACTOR)

MAIN_HEADING_POSITION = round(HEIGHT*0.1)
MAIN_HEADING_POSITION_OFFSET_X = round(WIDTH * 0.028288543140028)
PRIZE_NUMBER_TEXT = 'PRIZE NUMBER'
PRIZE_NUMBER_TEXT_POSITION = round(HEIGHT*0.2)
PRIZE_POSITION = round(HEIGHT*0.28)
TOTAL_PRIZE = 300
DESTINATION_DIR = '../qr_codes/'
FONT_NORMAL = '../assets/fonts/calibri.ttf'
FONT_BOLD = '../assets/fonts/calibrib.ttf'
QR_CODE_OFFSET_Y = round(HEIGHT*0.15)
CAP_IMAGE = Image.open('../assets/images/santa_cap.png')
CAP_IMAGE_WIDTH, CAP_IMAGE_HEIGHT = (round(WIDTH*0.10608203677511), round(HEIGHT*0.15))
CAP_IMAGE = CAP_IMAGE.rotate(45)
CAP_IMAGE = CAP_IMAGE.resize((CAP_IMAGE_WIDTH, CAP_IMAGE_HEIGHT))
#CAP_IMAGE = CAP_IMAGE.transform((CAP_IMAGE_WIDTH, CAP_IMAGE_HEIGHT), Image.EXTENT, CAP_IMAGE)
CHRISTMAS_TREE_IMAGE = Image.open('../assets/images/christmas_tree.jpg')
CHRISTMAS_TREE_IMAGE_WIDTH, CHRISTMAS_TREE_IMAGE_HEIGHT = (round(WIDTH*0.21216407355021), round(HEIGHT*0.40))
CHRISTMAS_TREE_IMAGE = CHRISTMAS_TREE_IMAGE.resize((CHRISTMAS_TREE_IMAGE_WIDTH, CHRISTMAS_TREE_IMAGE_HEIGHT))

CHRISTMAS_BALLS = Image.open('../assets/images/christmas_ball.jpg')
CHRISTMAS_BALLS_WIDTH, CHRISTMAS_BALLS_HEIGHT = (round(WIDTH*0.14144271570014), round(HEIGHT*0.40))
CHRISTMAS_BALLS = CHRISTMAS_BALLS.resize((CHRISTMAS_BALLS_WIDTH, CHRISTMAS_BALLS_HEIGHT))

for i in range(1, (SHEET.max_row)):
    img_original = Image.new('RGBA', (WIDTH, HEIGHT), color = (255,255,255, 255))
    fnt = ImageFont.truetype(FONT_BOLD, 30)
    d = ImageDraw.Draw(img_original)
    w, h = d.textsize(MAIN_HEADING, font=fnt)
    d.text(((WIDTH-w)/2 +MAIN_HEADING_POSITION_OFFSET_X, MAIN_HEADING_POSITION), MAIN_HEADING , font=fnt, fill=(0, 0, 0))
    fnt = ImageFont.truetype(FONT_NORMAL, 15)
    w, h = d.textsize(PRIZE_NUMBER_TEXT, font=fnt)
    d.text(((WIDTH-w)/2, PRIZE_NUMBER_TEXT_POSITION), PRIZE_NUMBER_TEXT , font=fnt, fill=(0, 0, 0))
    fnt = ImageFont.truetype(FONT_BOLD, 50)
    w, h = d.textsize(str(i), font=fnt)
    d.text(((WIDTH-w)/2, PRIZE_POSITION), str(i) , font=fnt, fill=(0, 0, 0))
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=5,
        border=0,
    )
    qr.add_data({'prize_number' : i, 'year': YEAR, 'prize_name' : SHEET['{}{}'.format(PRIZE_COULUMN, i+1)].value})
    qr.make(fit=True)

    img = qr.make_image(fill_color="black", back_color="white")
    w = img.pixel_size
    img_original.paste(CAP_IMAGE, (round(WIDTH*0.028288543140028), round(MAIN_HEADING_POSITION*.75)), mask=CAP_IMAGE)
    img_original.paste(CHRISTMAS_TREE_IMAGE, ((WIDTH-CHRISTMAS_TREE_IMAGE_WIDTH), (HEIGHT-CHRISTMAS_TREE_IMAGE_HEIGHT-10)))
    img_original.paste(CHRISTMAS_BALLS, (round(WIDTH*0.028288543140028), round(HEIGHT*0.25)))
    img_original.paste(img.get_image(), (round((WIDTH-w)/2), round((HEIGHT-w)/2)+QR_CODE_OFFSET_Y))
    img_original = ImageOps.expand(img_original, border=(2,2,2,2), fill=(0,0,0))
    img_original.save('{}Prize_Number_{}.png'.format(DESTINATION_DIR, i))