from pymongo import MongoClient
import json
from openpyxl import load_workbook
from tqdm import tqdm

with open('credentials.json') as fptr:
    data = json.load(fptr)
    
prize_list_file = "..\qr_code_generator\prize_list\PrizeList.xlsx"
name_list_file = "Name.xlsx"

def get_upload_data(file_name):
    heading = []
    workbook = load_workbook(file_name)
    sheet = workbook.active
    return_array = []
    for row_index, row in enumerate(sheet.iter_rows(1, sheet.max_row)):
        temp_dict = {}
        for column_index, each_cell in enumerate(row):
            if row_index == 0:
                heading.append(each_cell.value)
            else:
                temp_dict[heading[column_index]] = each_cell.value
        if temp_dict:
            return_array.append(temp_dict)
    return return_array
            
def update_database(collection, list_to_update):
    print("Updating {}".format(collection.name))
    for index, each_object in tqdm(enumerate(list_to_update)):
        result = collection.insert_one(each_object)
        if not result.acknowledged:
            print('Error in Creating ', each_object)

def get_database():
    CONNECTION_STRING = 'mongodb://{}:{}@localhost:27017/'.format(data['username'], data['password'])
    client = MongoClient(CONNECTION_STRING)
    return client['Prize2023']

if __name__ == "__main__":
    dbname = get_database()
    collection = dbname['PrizeList']
    prize_data_to_upload = get_upload_data(prize_list_file)
    update_database(collection, prize_data_to_upload)
    collection = dbname['UserData']
    name_data_to_upload = get_upload_data(name_list_file)
    update_database(collection, name_data_to_upload)
    